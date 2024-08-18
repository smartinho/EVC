from datetime import datetime, date
import openpyxl
from openpyxl.styles import PatternFill, Border, Font, Alignment, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import dataframe

def save_today(df_data, model_name):
    # 오늘 날짜를 YYYYMMDD 형식으로 얻기
    today_date = datetime.today().strftime('%Y%m%d')
    file_name = model_name + f'_{today_date}.xlsx'
    sheet_name = file_name[:-5]

    # 데이터프레임을 엑셀 파일로 저장
    df_data.to_excel(file_name, sheet_name = sheet_name, index=False)
    return file_name

def auto_fit_columns(file_name):
    wb = openpyxl.load_workbook(file_name)

    # 숫자 형식의 셀에 천 단위 쉼표 추가
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

    if ws['G1'].value is None:     
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'

            # 서식 적용 함수 호출
            apply_formatting(ws)

            # 각 열의 최대 길이에 맞춰 셀 너비 조정
            for col in ws.columns:
                max_length = 0
                column_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length + 2  # 추가 공간
                ws.column_dimensions[column_letter].width = adjusted_width

            # 'Spec.' 열의 너비 설정
            spec_column_index = None
            for col_idx in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col_idx).value == 'Spec.':
                    spec_column_index = col_idx
                    break
            if spec_column_index is not None:
                ws.column_dimensions[get_column_letter(spec_column_index)].width = 35

        # "G1" 셀에 환율 정보 삽입
        if not sheet_name == '미정단가':
            cell_value = f"환율[$] : {dataframe.rounded_usd_exchange_rate}원"
            ws['G1'] = cell_value

    wb.save(file_name)

def apply_formatting(ws):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    font_style = Font(name='Arial Narrow', size=11)
    light_gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    light_blue_fill = PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid")
    bold_font = Font(bold=True)

    ws.sheet_view.showGridLines = False

    if ws.title == '미정단가':
        max_cols = 4
    else:
        max_cols = 5
    # 모든 셀에 서식 적용
    for row in ws.iter_rows():
        for cell in row:
            if cell.column <= max_cols:
                cell.border = thin_border
                cell.font = font_style

    # 첫 번째 행에 옅은 회색 배경 및 볼드 적용
    for cell in ws[1]:
        if cell.column <= max_cols:
            cell.fill = light_gray_fill
            cell.font = bold_font
    
    target_column = 'A'
    target_column_index = openpyxl.utils.cell.column_index_from_string(target_column) - 1
    
    for row in ws.iter_rows():
        cell_value = str(row[target_column_index].value)  # target_column 열의 값을 가져옴
        
        for cell in row:
            if cell.value and ('Sum' in cell_value or 'Total' in cell_value):
                # 해당 행의 모든 셀에 배경색 적용
                for cell in row:
                    if cell.column <= max_cols:
                        cell.fill = light_blue_fill
                        cell.font = Font(bold=True)
                break  # 해당 행 전체에 색을 적용했으면 다음 행으로 이동

    # '구분' 열 값 병합
    merge_start = 1
    for i in range(2, ws.max_row + 1):
        if ws.cell(row=i, column=1).value != ws.cell(row=i - 1, column=1).value:
            if merge_start != i - 1:
                ws.merge_cells(start_row=merge_start, start_column=1, end_row=i - 1, end_column=1)
                ws.cell(row=merge_start, column=1).alignment = Alignment(vertical='top')
            merge_start = i
    if merge_start != ws.max_row:
        ws.merge_cells(start_row=merge_start, start_column=1, end_row=ws.max_row, end_column=1)
        ws.cell(row=merge_start, column=1).alignment = Alignment(vertical='top')

def copy_cell(source_cell, target_cell):
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell.font = source_cell.font.copy()
        target_cell.border = source_cell.border.copy()
        target_cell.fill = source_cell.fill.copy()
        target_cell.number_format = source_cell.number_format
        target_cell.protection = source_cell.protection.copy()
        target_cell.alignment = source_cell.alignment.copy()

def tbd_cost(selected_file):
    df = pd.read_excel(selected_file, skiprows=19)
    start_col_idx = df.columns.get_loc("Model.Suffix")
    df = df.iloc[:, start_col_idx:]

    wb = openpyxl.load_workbook(selected_file)
    ws = wb.active
    start_row = 20
    red_rows = []

    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            if cell.font.color and cell.font.color.rgb == 'FFFF0000':
                red_rows.append(row)
                break

    today_date = datetime.today().strftime('%Y%m%d')
    target_file = dataframe.model_name + f'_{today_date}.xlsx'
    target_sheet_name = '미정단가'

    target_wb = openpyxl.load_workbook(target_file)
    if target_sheet_name in target_wb.sheetnames:
        del target_wb[target_sheet_name]

    target_ws = target_wb.create_sheet(title=target_sheet_name)
    header_row = [cell for cell in ws[1]]

    for col_idx, source_cell in enumerate(header_row, start=1):
        target_cell = target_ws.cell(row=1, column=col_idx)
        copy_cell(source_cell, target_cell)

    red_rows.sort(key=lambda row: [cell.value if cell.value else '' for cell in row])

    for row_idx, row in enumerate(red_rows, start=2):
        for col_idx, source_cell in enumerate(row, start=1):
            target_cell = target_ws.cell(row=row_idx, column=col_idx)
            if isinstance(source_cell, openpyxl.cell.cell.MergedCell):
                source_cell = ws.cell(row=source_cell.row, column=source_cell.column)
            copy_cell(source_cell, target_cell)

    columns_to_delete = list(range(1, 8))
    columns_to_delete.sort(reverse=True)
    for col in columns_to_delete:
        target_ws.delete_cols(col, amount=1)

    target_wb.save(target_file)

    df_excel = pd.read_excel(target_file, sheet_name='미정단가', header=0)
    df_excel.columns = df.columns.tolist()
    df_excel = df_excel[['Part No', 'Desc.', 'Spec.', 'Exchange Rate(USD)']]

    with pd.ExcelWriter(target_file, engine='openpyxl', mode='a') as writer:
        writer.book.remove(writer.book[target_sheet_name])
        df_excel.to_excel(writer, sheet_name=target_sheet_name, index=False)

    auto_fit_columns(target_file)

