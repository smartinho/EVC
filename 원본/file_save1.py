from datetime import datetime
from datetime import date
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import pandas as pd

def Save_Today(df_data):
    # 오늘 날짜 얻기
    today_date = datetime.today().strftime('%Y%m%d')

    # 파일명 생성
    file_name = f'정리_{today_date}.xlsx'

    # 데이터프레임을 엑셀 파일로 저장
    df_data.to_excel(file_name, index=False)
    # print(file_name)
    return file_name

# 엑셀 파일 열기/서식적용/저장
def auto_fit_columns(file_name):
    wb = openpyxl.load_workbook(file_name)
    # ws = wb.active

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 셀의 숫자 형식을 3자리 단위로 쉼표를 넣어 포맷
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00' 

        # 서식 적용
        apply_formatting(wb, ws)
        
        # Iterate over all columns in the worksheet
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name

            # Find the maximum length of the data in the column
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            # Set the column width to the maximum length
            adjusted_width = max_length + 2  # Adding extra space
            ws.column_dimensions[column].width = adjusted_width

        # 'Spec.' 열 너비 설정
        column_width = 35
        spec_column = 'Spec.'  # Spec. 열의 헤더명

        # 'Spec.' 열의 열 번호 가져오기
        spec_column_index = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == spec_column:
                spec_column_index = col
                break
        
        # 열 너비 설정
        if spec_column_index is not None:
            ws.column_dimensions[openpyxl.utils.get_column_letter(spec_column_index)].width = column_width
            # print(f"'{spec_column}' 열의 너비를 {column_width}로 설정했습니다.")
        else:
            print(f"'{spec_column}' 열을 찾을 수 없습니다.")

    for sheet_name in pd.ExcelFile(file_name).sheet_names:
        if not sheet_name == '미정단가':
            # 시트 이름 변경
            today = date.today()
            today_str = today.strftime("%Y%m%d")
            new_sheet_name = today_str
            wb.active.title = new_sheet_name    
    
    # Save the modified workbook
    wb.save(file_name)

# 서식 적용
def apply_formatting(wb, ws):
    # Define the border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define the font style
    font_style = Font(name='Arial Narrow', size=11)

    # 눈금선 비활성화
    ws.sheet_view.showGridLines = False
    
    # Iterate over all cells in the worksheet
    for row in ws.iter_rows():
        for cell in row:
            # Apply the border to each cell
            cell.border = thin_border
            # Apply the font style to each cell
            cell.font = font_style

    # 옅은 회색 배경색 설정
    light_gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    # 옅은 푸른색 배경색 설정
    light_blue_fill = PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid")

    # bold 설정
    bold_font = Font(bold=True)

    # 컬럼의 모든 셀에 옅은 회색 및 Bold 적용
    row_number = 1  # 두 번째 행
    for cell in ws[row_number]:
        cell.fill = light_gray_fill
        cell.font = bold_font
    

    # 특정 열 기준 ('A' 열 예시, 열 인덱스는 0부터 시작)
    target_column = 'A'
    target_column_index = openpyxl.utils.cell.column_index_from_string(target_column) - 1

    # 엑셀의 모든 행을 순회
    for row in ws.iter_rows():
        cell_value = str(row[target_column_index].value)  # target_column 열의 값을 가져옴
        # if 'Sum' in cell_value or 'Total' in cell_value:
        for cell in row:
            if cell.value and ('Sum' in cell_value or 'Total' in cell_value):
                # 해당 행의 모든 셀에 배경색 적용
                for cell in row:
                    cell.fill = light_blue_fill
                    cell.font = Font(bold=True)
                break  # 해당 행 전체에 색을 적용했으면 다음 행으로 이동

    # '구분' 열의 값 병합
    merge_start = 1  # 초기 병합 시작 위치 (헤더를 제외한 첫 번째 행)
    for i in range(2, ws.max_row + 1):
        if ws[f'A{i}'].value != ws[f'A{i-1}'].value:
            if merge_start != i - 1:
                ws.merge_cells(start_row=merge_start, start_column=1, end_row=i - 1, end_column=1)
                ws[f'A{merge_start}'].alignment = ws[f'A{merge_start}'].alignment.copy(vertical='top')
            merge_start = i
    # 마지막 그룹 병합
    if merge_start != ws.max_row:
        ws.merge_cells(start_row=merge_start, start_column=1, end_row=ws.max_row, end_column=1)
        ws[f'A{merge_start}'].alignment = ws[f'A{merge_start}'].alignment.copy(vertical='top')

    # '구분' 열의 값 가운데 정렬
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

# Copy cell 서식 유지
def copy_cell(source_cell, target_cell):
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def copy(style):
    if style is None:
        return None
    return style.copy()

# 미정단가 정리
def TBD_Cost(selected_file):

   # 컬럼을 가져오기 위한 엑셀 파일 열기
    df = pd.read_excel(selected_file, skiprows=19)
    col_index = df.columns.get_loc("Model.Suffix")

    # 해당 열의 인덱스 이후의 열들만 선택
    df = df.iloc[:, col_index:]
    cols = df.columns.tolist()

    # 선택 파일 읽기
    wb = openpyxl.load_workbook(selected_file)
    ws = wb.active

    # 특정 행 번호 (예: 20번째 행부터)
    start_row = 20

    # 붉은색 글자의 행을 저장할 리스트
    red_rows = []

    # 첫 번째 행 (컬럼 헤더) 복사
    header_row = [cell for cell in ws[1]]

    # 특정 행부터 데이터를 복사
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            if cell.font.color and cell.font.color.rgb == 'FFFF0000':  # RGB 값이 FF0000 (붉은색)인지 확인
                red_rows.append(row)
                break
    
    today_date = datetime.today().strftime('%Y%m%d')
    target_file = f'정리_{today_date}.xlsx'
    target_sheet_name = '미정단가'
 
    target_wb = openpyxl.load_workbook(target_file)

    # 대상 엑셀 파일에 "미정단가" 시트가 있는지 확인
    if target_sheet_name in target_wb.sheetnames:
        # "미정단가" 시트가 있으면 삭제
        del target_wb[target_sheet_name]

    # "미정단가" 시트 추가
    target_ws = target_wb.create_sheet(title=target_sheet_name)

    #첫 번째 행 (컬럼 헤더) 추가
    for col_idx, source_cell in enumerate(header_row, start=1):
        target_cell = target_ws.cell(row=1, column=col_idx)
        copy_cell(source_cell, target_cell)

    # 붉은색 글자의 행을 정렬 (행 전체를 기준으로 정렬)
    red_rows.sort(key=lambda row: [cell.value if cell.value is not None else '' for cell in row])

    #정렬된 행을 새로운 워크북에 추가
    for row_idx, row in enumerate(red_rows, start=2):  # 두 번째 행부터 시작
        for col_idx, source_cell in enumerate(row, start=1):
            target_cell = target_ws.cell(row=row_idx, column=col_idx)
            if isinstance(source_cell, openpyxl.cell.cell.MergedCell):
                # 병합된 셀의 경우 대표 셀의 값을 사용
                source_cell = ws.cell(row=source_cell.row, column=source_cell.column)
            copy_cell(source_cell, target_cell)

    # 0~6열 삭제 (1-based index로 1~7열)
    columns_to_delete = list(range(1, 8))

    # 열 번호를 내림차순으로 정렬 (뒤에서부터 삭제해야 인덱스가 꼬이지 않음)
    columns_to_delete.sort(reverse=True)

    # 열 삭제 (첫 번째 행은 제외)
    for col in columns_to_delete:
        target_ws.delete_cols(col, amount=1)   
        
    # 파일 저장
    target_wb.save(target_file)

    # "미정단가" 시트를 읽고 첫 행을 컬럼명으로 설정
    df_excel = pd.read_excel(target_file, sheet_name='미정단가', header=0)
    df_excel.columns = cols
    df_excel=df_excel[['Part No','Desc.', 'Spec.', 'Exchange Rate(USD)']]
    # print(df_excel.head(5))
    with pd.ExcelWriter(target_file, engine='openpyxl', mode='a') as writer:
        # 기존 "미정단가" 시트를 삭제
        writer.book.remove(writer.book['미정단가'])
        # 변경된 데이터프레임을 "미정단가" 시트에 다시 쓰기
        df_excel.to_excel(writer, sheet_name='미정단가', index=False)

    auto_fit_columns(target_file)
